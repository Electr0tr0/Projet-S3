from psycopg2 import *
import openpyxl

# Connection parameters, modify as needed
parameters = "dbname=IKEA user=etu port=5434"

# Open a connection
try:
    connexion = connect(parameters)
    print("Connecté à la base !")
except Error as e:
    print(f"Error: {e}")
    exit(1)

# Create a cursor for running queries
curseur = connexion.cursor()

def getData(requete, listeParametres=()):
    try:
        curseur.execute(requete, listeParametres)
        return curseur.fetchall()
    except Error as e:
        print(f"Error: {e}")
        return []

def menuprincipal():
    requete = []
    a = int(input("Que voulez vous faire ?\n1-Rechercher des données sur une commande existante. \n2-Passer une nouvelle commande\n3-Avoir des renseignements sur des meubles\n"))
    if a == 1:
        b = input("Quel est le numéro de la commande désirée ?\n")
        print(menu1(b))
        export_to_excel(b,menu1(b),'menu1')
        
    elif a == 2:
        hasfinished = False
        while hasfinished==0 :
            c = int(input("Dans quelle pièce de la maison, le meuble sera-t'il installé ?\n1-Salon\n2-Salle à manger\n3-Cuisine\n4-Chambre\n5-Salle de Bain\n6-Bureau\n7-Entrée\n8-Jardin,Terrase et Balcon\n9-Buanderie et garage\n"))
            requete = requete + menu2(c)
            hasfinished=input("Avez vous finit ? 1 oui, 0 non")
        export_to_excel(c,requete,'menu2')
        
    elif a == 3:
        d = input("Quel type de renseignement voulez vous ?\n")
        menu3(d)
        export_to_excel(d,menu3(d),'menu3')
        
    else:
        print("La valeur rentrée ne correspond à aucune fonction")

def menu1(b):
    requete = getData("select tnom,pnom,mobid from synthese s where lower(tnom) like %s", (f'%{b.lower()}%',))
    return requete
    #export_to_excel(requete, 'menu1')

def menu2(c):
    e = c
    type = input("Quel type de meuble souhaitez-vous acheter ?\n")
    requete = getData("select tnom,pnom,mobid from synthese3 s where pieceid=%s and tnom like %s", (e, f'%{type}%'))
    return requete
    #export_to_excel(requete, 'menu2')

def menu3(d):
    f = d
    type = input("Dans quelle pièce ?\n")
    meuble = input("Quel type de meuble ?\n")
    requete = getData("select %s from synthese3 s where pieceid=%s and tnom like %s", (f,type, f'%{meuble}%'))
    return requete
    #export_to_excel(requete, 'menu2')


def export_to_excel(requete_utilisateur,data, sheet_name):
    # Create a workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    for i in range(len(requete_utilisateur.split(","))):
            sheet.cell(row=1,column=i,value=requete_utilisateur.split(",")[i])

    # Write the data to the sheet
    for i, row in enumerate(data, start=2):
        sheet.cell(row=i, column=1, value=row[0])
        sheet.cell(row=i, column=2, value=row[1])
        sheet.cell(row=i, column=3, value=row[2])

    # Save the workbook to a file
    workbook.save('Fichier.xlsx')
    print(f"Les données ont bien étées exportées à la page {sheet_name} dans Fichier.xlsx")

menuprincipal()

# Close the cursor and connection
curseur.close()
connexion.close()
#print("La connexion est fermée.")
